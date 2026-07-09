"""Convert a DPD Station export (STARTEC_U4*.xlsx) to a 26-column CSV
importable into Station.NET's address book ("carnet d'adresses").

Same field layout as build_station_import.py (26 cols, positions 0-25), but
with comma separator instead of tab so DPD Station's CSV import accepts it.

Dedupes on (NOM, CP, VILLE, ADRESSE). Skips rows already present in the
target file if it exists (append mode).

Usage:
    python build_station_addresses_csv.py <input.xlsx|csv> [output.csv]
"""

import csv
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

COUNTRY_MAP = {
    "FRANCE": "F", "PORTUGAL": "P", "POLOGNE": "PL", "BELGIQUE": "B",
    "ESPAGNE": "E", "ITALIE": "I", "ALLEMAGNE": "D", "LUXEMBOURG": "L",
    "SUISSE": "CH", "GRECE": "GR", "AUTRICHE": "A", "PAYS-BAS": "NL",
    "ROYAUME-UNI": "GB",
}

NUM_RE = re.compile(r"^(\d+\s*[A-Za-z]?)[,\s]+(.+)$")


def strip_accents(s):
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def parse_address(raw):
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    m = NUM_RE.match(raw)
    if m:
        return m.group(1).strip().rstrip(","), m.group(2).strip()
    return "", raw


def normalize_phone(raw):
    if not raw:
        return ""
    p = re.sub(r"\s+", "", str(raw).strip())
    if p.startswith("+33") and len(p) >= 11:
        p = "0" + p[3:]
    return p


def cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_existing(out_path: Path) -> set:
    if not out_path.exists():
        return set()
    keys = set()
    with out_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        for row in csv.reader(f):
            if len(row) < 14 or not row[0]:
                continue
            nom = row[0].upper()
            cp = row[9]
            ville = row[10].upper()
            num = row[13].strip()
            voie = row[12].strip()
            addr = (f"{num} {voie}".strip() if num else voie).upper()
            keys.add((nom, cp, ville, addr))
    return keys


def read_rows(in_path: Path):
    if in_path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(in_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [cell(c) for c in next(it)]
        for raw in it:
            if not any(cell(c) for c in raw):
                continue
            yield {h: cell(v) for h, v in zip(header, raw)}
    else:
        with in_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
            yield from csv.DictReader(f)


def main(in_path: Path, out_path: Path) -> int:
    existing = load_existing(out_path)
    print(f"Clients déjà dans {out_path.name}: {len(existing)}")

    seen = {}
    for row in read_rows(in_path):
        name = strip_accents(row.get("Nom du destinataire", "")).strip()
        cp = (row.get("Code postal") or "").strip()
        ville = strip_accents(row.get("Ville", "")).strip()
        adresse = strip_accents(row.get("Adresse de livraison", "")).strip()
        pays_raw = (row.get("Pays de destination") or "").strip().upper()
        tel = normalize_phone(row.get("Téléphone", "") or row.get("Telephone", ""))

        if not name or not cp:
            continue

        # CP français à 4 chiffres → padding 0
        if cp.isdigit() and len(cp) == 4 and pays_raw in ("FRANCE", ""):
            cp = "0" + cp

        key = (name.upper(), cp, ville.upper(), adresse.upper())
        if key in existing:
            continue
        prev = seen.get(key)
        if prev and prev["tel"]:
            continue
        if prev and not tel:
            continue

        num, voie = parse_address(adresse)
        pays = COUNTRY_MAP.get(pays_raw, pays_raw[:2] if pays_raw else "F")

        seen[key] = {
            "raison": name.upper()[:35],
            "commercial": name.title()[:35],
            "pays": pays,
            "cp": cp[:10],
            "ville": ville.upper()[:35],
            "voie": voie[:35],
            "num": num[:10],
            "tel": tel[:15],
        }

    mode = "a" if existing else "w"
    with out_path.open(mode, encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        for rec in seen.values():
            cols = [""] * 26
            cols[0] = rec["raison"]
            cols[2] = rec["commercial"]
            cols[8] = rec["pays"]
            cols[9] = rec["cp"]
            cols[10] = rec["ville"]
            cols[12] = rec["voie"]
            cols[13] = rec["num"]
            cols[15] = rec["tel"]
            w.writerow(cols)

    verb = "Append" if existing else "Wrote"
    print(f"{verb}: {len(seen)} client(s) unique(s) -> {out_path}")
    print(f"Total dans le fichier: {len(existing) + len(seen)}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1])
    if len(sys.argv) > 2:
        out_path = Path(sys.argv[2])
    else:
        stamp = datetime.now().strftime("%d%m%Y-%H%M%S")
        out_path = in_path.with_name(f"STARTEC_addresses_{stamp}.csv")
    sys.exit(main(in_path, out_path))
