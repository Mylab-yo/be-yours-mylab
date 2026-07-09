"""Split a Choose orders export into a DPD Station .DAT (Predict) and a Mondial Relay CSV.

Accepts a CSV or XLSX file. For XLSX: reads the "général" sheet and skips any row
where the "État" column equals "Terminé" (commandes déjà expédiées).

Choose columns: Référence, Commandé le, Prénom, Nom, N° et voie,
Instructions de livraison, Lieu dit, Code postal, Commune, Code ISO du pays,
Téléphone portable, ID Point de retrait [, État].

Routing rule:
  - ID Point de retrait empty  -> DPD Predict .DAT
  - ID Point de retrait present -> Mondial Relay CSV (mode 24R, V3.1 spec)

Output:
  - <out_dir>/DPDFRANCE_<DDMMYYYY>-<HHMMSS>.dat   (fixed-width 2247-char, $VERSION=110)
  - <out_dir>/MondialRelay_<DDMMYYYY>-<HHMMSS>.csv (Connect V3.1, 44 cols, ; cp1252)

Usage:
    python build_dpd_mr_from_choose.py <choose.csv|.xlsx> [out_dir]
"""

import csv
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Sender configuration (STARTEC / MY.LAB) — taken from the existing DPD export.
# ---------------------------------------------------------------------------
SENDER_ACCOUNT = "1951"
# Service code DPD (visible dans la liste déroulante Station) :
#   3043 = STARTEC (standard)
#   3044 = STARTEC Predict  ← utilisé pour tous les imports Choose (livraison à domicile + SMS)
#   3045 = STARTEC R
DPD_SERVICE_CODE = "00003044"
SENDER_NAME = "STARTEC"
SENDER_ZIP = "84300"
SENDER_CITY = "LES TAILLADES"
SENDER_ADDR = "245, ROUTE DE ROBION"
SENDER_COUNTRY = "F"
SENDER_PHONE = "+33 4 42 87 18 75"
SENDER_EMAIL = "hello@vegetal-origin.com"

# ---------------------------------------------------------------------------
# DPD Station .DAT fixed-width layout (offsets validated against a real export).
# Every record is 2247 chars, lines terminated with CRLF, encoded cp1252.
# ---------------------------------------------------------------------------
DAT_RECORD_LEN = 2247
DAT_VERSION_HEADER = "$VERSION=110"

# (offset, width, value_key) — value_key is looked up in the per-record dict.
# Anything not listed stays as space.
DAT_FIELDS = [
    (0,    4,  "sender_account"),
    (37,   23, "order_ref"),
    (60,   35, "recip_name1"),
    (95,   35, "recip_name2"),
    (270,  10, "recip_zip"),
    (280,  45, "recip_city"),
    (325,  45, "recip_addr"),
    (370,  1,  "recip_country"),
    (373,  45, "recip_phone"),
    (418,  35, "sender_name"),
    (628,  10, "sender_zip"),
    (638,  45, "sender_city"),
    (683,  45, "sender_addr"),
    (728,  1,  "sender_country"),
    (731,  45, "sender_phone"),
    (901,  10, "ship_date"),
    (911,  8,  "service_code"),
    (919,  4,  "sender_account"),
    (954,  4,  "sender_account"),
    (1035, 4,  "sender_account"),
    (1116, 80, "sender_email"),
    (1231, 80, "recip_email"),
    (1311, 35, "recip_mobile"),
    (1569, 35, "recip_lastname"),
]

# Choose ISO country → DPD single-letter country code (best-effort; F default).
DPD_COUNTRY_MAP = {
    "FR": "F", "BE": "B", "CH": "CH", "DE": "D", "ES": "E", "IT": "I",
    "LU": "L", "NL": "NL", "PT": "P", "GB": "GB", "UK": "GB", "MC": "F",
}

# ---------------------------------------------------------------------------
# Mondial Relay Connect — spec officielle V3.1 (doc "import-de-fichiers-csv-v3.1.pdf")
# 44 champs (A → AR) séparés par `;`, MAJUSCULES sans accents, 75 envois max/fichier.
# ---------------------------------------------------------------------------
MR_MODE_LIVRAISON = "24R"   # col S/19 — point relais 24h
MR_POIDS_DEFAUT_G = "500"   # col V/22 — placeholder grammes (3-7 chiffres requis)
MR_NAME_RE = re.compile(r"[^0-9A-Z_\-'., /]")    # cols C/E/D/F autorisés
MR_VILLE_RE = re.compile(r"[^A-Z_\-' ]")          # col G — pas de chiffres


# ---------------------------------------------------------------------------
def normalize_zip(raw, iso):
    z = (raw or "").strip()
    if iso == "FR" and z.isdigit() and len(z) == 4:
        return "0" + z
    return z


def normalize_phone(raw):
    """DPD Station veut le format national 10 chiffres (0XXXXXXXXX)."""
    if not raw:
        return ""
    p = raw.strip().lstrip("'").lstrip("=")
    p = re.sub(r"[\s\.\-\(\)/]+", "", p)
    if p.startswith("+33") and len(p) >= 11:
        p = "0" + p[3:]
    elif p.startswith("0033") and len(p) >= 12:
        p = "0" + p[4:]
    return p


def phone_international(raw, iso="FR"):
    """MR Connect FR : `0XXXXXXXXX` (10c) ou `+33XXXXXXXXX` ou `0033XXXXXXXXX`."""
    if not raw:
        return ""
    p = raw.strip().lstrip("'").lstrip("=")
    p = re.sub(r"[\s\.\-\(\)/]+", "", p)
    if p.startswith("+"):
        return p
    if p.startswith("0033"):
        return "+" + p[2:]
    if iso == "FR" and p.startswith("0") and len(p) == 10:
        return "+33" + p[1:]
    return p


def strip_accents(s):
    """é → e, à → a, ï → i, ç → c, etc. Garde l'apostrophe droite."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def mr_upper(s, max_len, allowed_re=MR_NAME_RE):
    """MR exige UPPERCASE sans accents, restreint au charset autorisé du champ."""
    cleaned = strip_accents(s or "").upper()
    cleaned = cleaned.replace("’", "'")  # apostrophe typographique → droite
    cleaned = allowed_re.sub(" ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:max_len]


def mr_ref(s, max_len=15):
    """Référence expédition : ^[0-9A-Z_ -]{0,15}$ — uppercase, alphanum + _ - espace."""
    cleaned = strip_accents(s or "").upper()
    cleaned = re.sub(r"[^0-9A-Z_ \-]", "", cleaned)
    return cleaned[:max_len]


def mr_relay_id(raw):
    """ID Relais : 6 chiffres fixes. Choose envoie 5 chiffres → padding à 6."""
    digits = re.sub(r"\D", "", raw or "")
    if not digits:
        return ""
    return digits.zfill(6)[:6]


def truncate(value, width):
    s = "" if value is None else str(value)
    return s[:width]


def stamp(record, offset, width, value):
    s = truncate(value, width)
    return record[:offset] + s.ljust(width) + record[offset + width:]


def build_dat_record(fields):
    record = " " * DAT_RECORD_LEN
    for offset, width, key in DAT_FIELDS:
        record = stamp(record, offset, width, fields.get(key, ""))
    return record[:DAT_RECORD_LEN]


def short_ref(choose_ref, max_len=23):
    """Choose order refs are 25-char hex hashes. Trim to fit DPD's 23-char slot."""
    return (choose_ref or "")[:max_len]


def is_relay(row):
    return bool((row.get("ID Point de retrait") or "").strip())


def _xlsx_cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_choose_rows(in_path: Path):
    """Return Choose rows as list of dict. Drops rows where État == 'Terminé'."""
    if in_path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(in_path, data_only=True, read_only=True)
        sheet_name = "général" if "général" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        header = [(_xlsx_cell(c) or "") for c in next(it)]
        rows = []
        for raw in it:
            if not any(_xlsx_cell(c) for c in raw):
                continue
            row = {h: _xlsx_cell(v) for h, v in zip(header, raw)}
            etat = (row.get("État") or row.get("Etat") or "").strip().lower()
            if etat.startswith("termin"):
                continue
            nom = (row.get("Nom") or "").strip()
            prenom = (row.get("Prénom") or row.get("Prenom") or "").strip()
            cp = (row.get("Code postal") or "").strip()
            if not (nom or prenom) or not cp:
                continue
            rows.append(row)
        return rows
    with in_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        return list(csv.DictReader(f))


def main(in_path: Path, out_dir: Path) -> int:
    out_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    stamp_fname = now.strftime("%d%m%Y-%H%M%S")

    rows = load_choose_rows(in_path)
    dpd_rows = [r for r in rows if not is_relay(r)]
    mr_rows = [r for r in rows if is_relay(r)]

    # ---- DPD Station .DAT ----------------------------------------------------
    dat_lines = [DAT_VERSION_HEADER]
    for row in dpd_rows:
        first = (row.get("Prénom") or "").strip()
        last = (row.get("Nom") or "").strip()
        full = f"{first} {last}".strip()
        iso = (row.get("Code ISO du pays") or "FR").strip().upper()
        phone = normalize_phone(row.get("Téléphone portable", ""))
        fields = {
            "sender_account": SENDER_ACCOUNT,
            "order_ref": short_ref(row.get("Référence", "")),
            "recip_name1": full.upper(),
            "recip_name2": "",
            "recip_zip": normalize_zip(row.get("Code postal"), iso),
            "recip_city": (row.get("Commune") or "").strip(),
            "recip_addr": (row.get("N° et voie") or "").strip(),
            "recip_country": DPD_COUNTRY_MAP.get(iso, "F"),
            "recip_phone": phone,
            "sender_name": SENDER_NAME,
            "sender_zip": SENDER_ZIP,
            "sender_city": SENDER_CITY,
            "sender_addr": SENDER_ADDR,
            "sender_country": SENDER_COUNTRY,
            "sender_phone": SENDER_PHONE,
            "ship_date": now.strftime("%d/%m/%Y"),
            "service_code": DPD_SERVICE_CODE,
            "sender_email": SENDER_EMAIL,
            "recip_email": "",
            "recip_mobile": phone,
            "recip_lastname": last.upper(),
        }
        dat_lines.append(build_dat_record(fields))

    dat_path = out_dir / f"DPDFRANCE_{stamp_fname}.dat"
    with dat_path.open("wb") as f:
        for line in dat_lines:
            f.write(line.encode("cp1252", errors="replace") + b"\r\n")
    print(f"[DPD]  {dat_path.name}: {len(dpd_rows)} commande(s) Predict")

    # ---- Mondial Relay Connect — 44 colonnes (spec officielle V3.1) ---------
    # Pas de header. MAJUSCULES sans accents. Cols A→AR (43 ';' par ligne).
    mr_path = out_dir / f"MondialRelay_{stamp_fname}.csv"
    with mr_path.open("w", encoding="cp1252", errors="replace", newline="") as f:
        w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        for row in mr_rows:
            first = (row.get("Prénom") or "").strip()
            last = (row.get("Nom") or "").strip()
            iso = (row.get("Code ISO du pays") or "FR").strip().upper()
            tel = phone_international(row.get("Téléphone portable", ""), iso)
            relay_id = mr_relay_id(row.get("ID Point de retrait"))
            ref = mr_ref(row.get("Référence", ""), 15)
            nom_complet = mr_upper(f"{last} {first}", 32)
            addr1 = mr_upper(row.get("N° et voie", ""), 32)
            addr2 = mr_upper(row.get("Lieu dit", ""), 32)
            ville = mr_upper(row.get("Commune", ""), 25, MR_VILLE_RE)
            cp = normalize_zip(row.get("Code postal"), iso)
            w.writerow([
                "",                       # A  1  N° de Client
                ref,                      # B  2  Référence expédition
                nom_complet,              # C  3  NOM PRENOM (livraison)
                "",                       # D  4  Complément du nom
                addr1,                    # E  5  Numéro + Rue
                addr2,                    # F  6  Complément d'adresse
                ville,                    # G  7  Ville
                cp,                       # H  8  Code Postal
                iso,                      # I  9  Pays destinataire
                "",                       # J 10  Téléphone fixe
                tel,                      # K 11  Téléphone cellulaire
                "",                       # L 12  Email
                "A",                      # M 13  Type Collecte (A = Agence)
                "",                       # N 14  ID Relais Collecte
                "",                       # O 15  Code Pays Collecte
                "R",                      # P 16  Type Livraison (R = Relais)
                relay_id,                 # Q 17  ID Relais Livraison (6 chiffres)
                iso,                      # R 18  Code Pays Relais Livraison
                MR_MODE_LIVRAISON,        # S 19  Mode Livraison (24R)
                "FR",                     # T 20  Langue Destinataire
                "1",                      # U 21  Nombre colis
                MR_POIDS_DEFAUT_G,        # V 22  Poids (g, 3-7 chiffres)
                "",                       # W 23  Longueur
                "",                       # X 24  Volume
                "",                       # Y 25  Valeur expédition
                "",                       # Z 26  Devise
                "",                       # AA 27 Assurances
                "",                       # AB 28 Montant CRT
                "",                       # AC 29 Devise CRT
                "",                       # AD 30 Instructions Livraison
                "",                       # AE 31 Top Avisage
                "",                       # AF 32 Top Reprise Domicile
                "",                       # AG 33 Temps Montage
                "",                       # AH 34 Top Rendez-vous (laisser vide)
                "",                       # AI 35 Article 01
                "",                       # AJ 36 Article 02
                "",                       # AK 37 Article 03
                "",                       # AL 38 Article 04
                "",                       # AM 39 Article 05
                "",                       # AN 40 Article 06
                "",                       # AO 41 Article 07
                "",                       # AP 42 Article 08
                "",                       # AQ 43 Article 09
                "",                       # AR 44 Article 10
            ])
    print(f"[MR]   {mr_path.name}: {len(mr_rows)} commande(s) point relais")
    print(f"Total: {len(rows)} commandes -> DPD={len(dpd_rows)} / MR={len(mr_rows)}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.parent
    sys.exit(main(in_path, out_dir))
