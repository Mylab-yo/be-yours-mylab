"""One-off: convert STARTEC_U4.xlsx -> CSV for build_station_import.py."""
import csv
import sys
from pathlib import Path
import openpyxl

src = Path(sys.argv[1])
dst = Path(sys.argv[2])

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb.active

with dst.open("w", encoding="utf-8", newline="") as f:
    writer = csv.writer(f)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        out_row = []
        for cell in row:
            if cell is None:
                out_row.append("")
            elif hasattr(cell, "strftime"):  # datetime
                out_row.append(cell.strftime("%Y-%m-%d"))
            else:
                out_row.append(str(cell))
        writer.writerow(out_row)

print(f"Wrote {dst} ({ws.max_row} rows, {ws.max_column} cols)")
