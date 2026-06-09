"""Convert a DPD Station export (STARTEC_U4*.xlsx) back into an importable .DAT.

The export from DPD Station has 25 columns (No colis, Service, Adresse, etc.).
This script keeps the recipient info and rebuilds a fresh .DAT in the same
fixed-width format as build_dpd_mr_from_choose.py uses (2247 chars/ligne,
$VERSION=110, CRLF, cp1252).

Service code (col `No CPTE client`):
  3043 = STARTEC Classic  -> kept as-is
  3044 = STARTEC Predict
  3045 = STARTEC R

Usage:
    python convert_startec_u4_to_dat.py <input.xlsx> [output.dat]
"""

import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

# Reuse the same sender + field layout as build_dpd_mr_from_choose.py.
sys.path.insert(0, str(Path(__file__).parent))
from build_dpd_mr_from_choose import (
    DAT_FIELDS, DAT_RECORD_LEN, DAT_VERSION_HEADER, DPD_COUNTRY_MAP,
    SENDER_ACCOUNT, SENDER_ADDR, SENDER_CITY, SENDER_COUNTRY, SENDER_EMAIL,
    SENDER_NAME, SENDER_PHONE, SENDER_ZIP, build_dat_record, normalize_phone,
)


PAYS_MAP = {"FRANCE": "FR", "BELGIQUE": "BE", "ALLEMAGNE": "DE",
            "ESPAGNE": "ES", "ITALIE": "IT", "SUISSE": "CH",
            "LUXEMBOURG": "LU", "PAYS-BAS": "NL", "PORTUGAL": "PT"}


def strip_accents(s):
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main(in_path: Path, out_path: Path) -> int:
    wb = openpyxl.load_workbook(in_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [cell(c) for c in next(it)]
    idx = {h: i for i, h in enumerate(header)}

    def col(row, name, default=""):
        i = idx.get(name)
        return cell(row[i]) if i is not None and i < len(row) else default

    now = datetime.now()
    dat_lines = [DAT_VERSION_HEADER]
    skipped = 0

    for row in it:
        if not any(cell(c) for c in row):
            continue
        nom = strip_accents(col(row, "Nom du destinataire")).upper()
        addr = strip_accents(col(row, "Adresse de livraison"))
        zip_ = col(row, "Code postal")
        ville = strip_accents(col(row, "Ville")).upper()
        tel = normalize_phone(col(row, "Téléphone"))
        pays = col(row, "Pays de destination").upper()
        iso = PAYS_MAP.get(pays, "FR")
        service_code = col(row, "No CPTE client") or SENDER_ACCOUNT.lstrip("0")
        ref = (col(row, "Votre référence 1") or col(row, "No colis"))[:23]

        if not (nom and addr and zip_ and ville):
            skipped += 1
            continue

        # CP français à 4 chiffres → padding 0
        if iso == "FR" and zip_.isdigit() and len(zip_) == 4:
            zip_ = "0" + zip_

        # Dernier mot du nom comme lastname (pour le slot [1569])
        tokens = nom.split()
        lastname = tokens[-1] if tokens else nom

        fields = {
            "sender_account": SENDER_ACCOUNT,
            "order_ref": ref,
            "recip_name1": nom,
            "recip_name2": "",
            "recip_zip": zip_,
            "recip_city": ville,
            "recip_addr": addr,
            "recip_country": DPD_COUNTRY_MAP.get(iso, "F"),
            "recip_phone": tel,
            "sender_name": SENDER_NAME,
            "sender_zip": SENDER_ZIP,
            "sender_city": SENDER_CITY,
            "sender_addr": SENDER_ADDR,
            "sender_country": SENDER_COUNTRY,
            "sender_phone": SENDER_PHONE,
            "ship_date": now.strftime("%d/%m/%Y"),
            "service_code": service_code.zfill(8),
            "sender_email": SENDER_EMAIL,
            "recip_email": "",
            "recip_mobile": tel,
            "recip_lastname": lastname,
        }
        dat_lines.append(build_dat_record(fields))

    with out_path.open("wb") as f:
        for line in dat_lines:
            f.write(line.encode("cp1252", errors="replace") + b"\r\n")
    print(f"[DPD] {out_path.name}: {len(dat_lines)-1} ligne(s) importables, {skipped} ligne(s) ignorée(s) (nom/adresse manquant)")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1])
    if len(sys.argv) > 2:
        out_path = Path(sys.argv[2])
    else:
        stamp = datetime.now().strftime("%d%m%Y-%H%M%S")
        out_path = in_path.with_name(f"DPDFRANCE_STARTEC_U4_{stamp}.dat")
    sys.exit(main(in_path, out_path))
